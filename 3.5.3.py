import numpy as np
import sqlite3
import pandas as pd
from openpyxl.styles import Side, Border, Font
from matplotlib import pyplot as plt
import pdfkit
from jinja2 import Environment, FileSystemLoader
from openpyxl import Workbook

class Statistics:
    def __init__(self, vacancy_name: str, area_name: str):
        self.vacancy_name = vacancy_name
        self.area_name = area_name
        self.vacancies_count = 0
        self.years_df = pd.DataFrame()
        self.cities_salary_df = pd.DataFrame()
        self.cities_percent_df = pd.DataFrame()

    def sort_dataframe(self, by: str) -> pd.DataFrame:
        return self.cities_salary_df.sort_values(by=by, ascending=False).head(10)

    def get_percent_of_other_cities(self) -> float:
        return 100 - self.cities_percent_df['percent'].sum()

    def get_statistics(self) -> None:
        connect = sqlite3.connect('vacancies_database.db')
        cursor = connect.cursor()
        self.vacancies_count = cursor.execute(self.get_sql_requests['count']).fetchone()[0]
        years_df = pd.read_sql(self.get_sql_requests['years'], connect, index_col='year')
        prof_years_df = pd.read_sql(self.get_sql_requests['prof_years'], connect, index_col='year')
        self.cities_salary_df = pd.read_sql(self.get_sql_requests['cities_salary'], connect, index_col='area_name').fillna(0)\
            .astype({'salary': np.int})
        self.years_df = years_df.join(prof_years_df).fillna(0) \
            .astype({'salary': np.int, 'count': np.int, 'prof_salary': np.int, 'prof_count': np.int})
        self.cities_percent_df = pd.read_sql(self.get_sql_requests['cities_percent'], connect, index_col='area_name').fillna(0)

    @property
    def get_sql_requests(self) -> dict:
        return {
            'years': 'SELECT CAST(SUBSTRING(published_at, 0, 5) as INTEGER) as year, '
                     'CAST(AVG(Salary) as INTEGER) as salary, '
                     'CAST(COUNT("index") as INTEGER) as count '
                     'FROM vacancies GROUP BY year',
            'prof_years': 'SELECT CAST(SUBSTRING(published_at, 0, 5) as INTEGER) as year, '
                          'CAST(AVG(Salary) as INTEGER) as prof_salary, '
                          'CAST(COUNT("index") as INTEGER) as prof_count '
                          'FROM vacancies '
                          f'WHERE name LIKE ("%{self.vacancy_name}%") '
                          f'AND area_name LIKE ("%{self.area_name}%") '
                          'GROUP BY year',
            'count': 'SELECT COUNT(*) FROM vacancies',
            'cities_salary': 'SELECT area_name, '
                             'CAST(AVG(Salary) as INTEGER) as salary '
                             'FROM vacancies '
                             'GROUP BY area_name '
                             'ORDER BY salary DESC '
                             'LIMIT 10',
            'cities_percent': 'SELECT area_name, '
                              f'(COUNT("index") * 100.0 / {self.vacancies_count}) as percent '
                              'FROM vacancies '
                              'GROUP BY area_name '
                              'HAVING percent > 1 '
                              'ORDER BY percent DESC '
                              'LIMIT 10'
        }

class ExcelCreator:
    def __init__(self, stats: Statistics):
        self.workbook = self.initialize_workbook(stats)

    def initialize_workbook(self, statistics: Statistics) -> Workbook:
        workbook, years_sheet, cities_sheet = self.create_workbook(statistics)
        self.add_stats_to_excel(statistics, years_sheet, cities_sheet)
        self.set_sheets_settings(statistics, years_sheet, cities_sheet)
        workbook.save('report.xlsx')
        return workbook

    def create_workbook(self, statistics: Statistics) -> tuple:
        workbook = Workbook()

        years_sheet = workbook.active
        years_sheet.title = 'Статистика по годам'
        cities_sheet = workbook.create_sheet('Статистика по городам')

        years_sheet.append(
            ['Год', 'Средняя зарплата', f'Средняя зарплата - {statistics.vacancy_name} по региону {statistics.area_name}',
             'Количество вакансий', f'Количество вакансий - {statistics.vacancy_name} по региону {statistics.area_name}'])
        cities_sheet.append(['Город', 'Уровень зарплат', '', 'Город', 'Доля Вакансий'])

        return workbook, years_sheet, cities_sheet

    def add_stats_to_excel(self, statistics: Statistics, years_sheet: Workbook.worksheets, cities_sheet: Workbook.worksheets):
        for year in statistics.years_df.index:
            years_sheet.append([year, statistics.years_df.loc[year]['salary'], statistics.years_df.loc[year]['prof_salary'],
                                statistics.years_df.loc[year]['count'], statistics.years_df.loc[year]['prof_count']])

        for city in statistics.cities_salary_df.index:
            cities_sheet.append([city, statistics.cities_salary_df.loc[city]['salary']])

        for i, city in enumerate(statistics.cities_percent_df.index, 2):
            cities_sheet[f'D{i}'].value = city
            cities_sheet[f'E{i}'].value = f'{round(statistics.cities_percent_df.loc[city]["percent"], 2)}%'

    def set_sheets_settings(self, stats: Statistics, years_sheet: Workbook.worksheets, cities_sheet: Workbook.worksheets) -> None:
        used_columns = ['A', 'B', 'C', 'D', 'E']
        for i in used_columns:
            years_sheet[f'{i}1'].font = Font(bold=True)
            cities_sheet[f'{i}1'].font = Font(bold=True)
            years_sheet.column_dimensions[i].width = max(map(lambda x: len(str(x.value)), years_sheet[i])) + 1
            cities_sheet.column_dimensions[i].width = max(
                map(lambda x: len(str(x.value)), cities_sheet[i])) + 1

        thins = Side(border_style="thin")
        for column in used_columns:
            for row in range(1, 12):
                years_sheet[f'{column}{row}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)

        for column in used_columns:
            for row in range(1, len(stats.years_df.index) + 2):
                if column == 'C':
                    break
                cities_sheet[f'{column}{row}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)

class GraphsCreator:
    def __init__(self, statistics):
        self.generate_image(statistics)

    def generate_image(self, statistics: Statistics) -> None:
        plt.subplots(figsize=(10, 7))
        plt.grid(True)

        self.create_salary_by_year_plot(statistics)
        self.create_count_by_year_plot(statistics)
        self.create_salary_by_city_plot(statistics)
        self.create_count_by_city_plot(statistics)

        plt.subplots_adjust(wspace=0.5, hspace=0.5)
        plt.savefig('graph.png')

    def create_salary_by_year_plot(self, statistics: Statistics) -> None:
        first = plt.subplot(221)
        plt.tick_params(axis='x', which='major', labelsize=8, rotation=90)
        plt.tick_params(axis='y', which='major', labelsize=8)
        first.bar(list(map(lambda y: y - 0.2, statistics.years_df.index)),
                  statistics.years_df['salary'], width=0.4,
                  label='Средняя з/п')
        first.bar(list(map(lambda y: y + 0.2, statistics.years_df.index)),
                  statistics.years_df['prof_salary'], width=0.4,
                  label=f'З/п {statistics.vacancy_name} по региону {statistics.area_name}')
        plt.legend(fontsize=8)
        plt.title('Уровень зарплат по годам', fontsize=12)

    def create_count_by_year_plot(self, statistics: Statistics) -> None:
        second = plt.subplot(222)
        plt.tick_params(axis='x', which='major', labelsize=8, rotation=90)
        plt.tick_params(axis='y', which='major', labelsize=8)
        second.bar(list(map(lambda y: y - 0.2, statistics.years_df.index)),
                   statistics.years_df['count'], width=0.4,
                   label='Количество вакансий')
        second.bar(list(map(lambda y: y + 0.2, statistics.years_df.index)),
                   statistics.years_df['prof_count'].to_list(), width=0.4,
                   label=f'Количество вакансий {statistics.vacancy_name} по региону {statistics.area_name}')
        plt.legend(fontsize=8)
        plt.title('Количество вакансий по годам', fontsize=12)

    def create_salary_by_city_plot(self, statistics: Statistics) -> None:
        third = plt.subplot(223)
        plt.tick_params(axis='x', which='major', labelsize=8)
        plt.tick_params(axis='y', which='major', labelsize=6)
        third.barh(statistics.cities_salary_df.index, statistics.cities_salary_df['salary'])
        plt.title('Уровень зарплат по городам', fontsize=12)

    def create_count_by_city_plot(self, statistics: Statistics) -> None:
        statistics.get_percent_of_other_cities()
        fourth = plt.subplot(224)
        plt.rc('xtick', labelsize=6)
        cities_df = statistics.cities_percent_df
        cities_df.loc['Другие'] = statistics.get_percent_of_other_cities()
        fourth.pie(cities_df['percent'],
                   labels=cities_df.index,
                   colors=['r', 'g', 'b', 'm', 'y', 'c', 'orange', 'darkblue', 'pink', 'sienna', 'grey'])
        plt.title('Доля вакансий по городам', fontsize=12)

class PdfCreator:
    def __init__(self, vacancy_name: str, area_name: str, workbook: Workbook, years_sheet_rows: int):
        config = pdfkit.configuration(wkhtmltopdf=r'D:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        options = {'enable-local-file-access': None}
        pdfkit.from_string(self.fill_pdf_template(vacancy_name, area_name, workbook["Статистика по годам"],
                                                  workbook["Статистика по городам"], years_sheet_rows),
                           'report.pdf', configuration=config, options=options)

    def fill_pdf_template(self, vacancy_name: str, area_name: str, years_sheet: Workbook.worksheets,
                          cities_sheet: Workbook.worksheets, years_sheet_rows: int) -> str:
        env = Environment(loader=FileSystemLoader('./ReportModule'))
        template = env.get_template('pdf_template.html')
        pdf_template = template.render({'vacancie_name': vacancy_name,
                                        'area_name': area_name,
                                        'years_table': self.create_html_table(years_sheet, years_sheet_rows),
                                        'cities_table_first': self.create_html_table(cities_sheet, 10, last_column=2),
                                        'cities_table_second': self.create_html_table(cities_sheet, 10, 4)})
        return pdf_template

    def create_html_table(self, ws: Workbook.worksheets, rows_count: int, first_column: int = 1,
                          last_column: int = 5) -> str:
        html = ''
        is_first = True
        for row in ws.iter_rows(min_row=1, min_col=first_column, max_col=last_column, max_row=rows_count + 1):
            html += '<tr>'
            for cell in row:
                html += '<td><b>' + str(cell.value) + '</b></td>' if is_first else '<td>' + str(cell.value) + '</td>'
            html += '</tr>'
            is_first = False
        return html

class Report:
    def __init__(self, vacancy_name: str, area_name: str):
        self.statistics = Statistics(vacancy_name, area_name)
        self.statistics.get_statistics()
        GraphsCreator(self.statistics)
        PdfCreator(vacancy_name, area_name, ExcelCreator(self.statistics).workbook, len(self.statistics.years_df.index))