import csv
import re
import datetime
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, Side, Border
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.utils import get_column_letter

class DataSet:
    def __init__(self, file):
        self.file = file
        self.vacancies = [Vacancy(vac) for vac in self.csv_filer(*self.csv_reader(file))]

    def delete_html(self, new_html):
        result = re.compile(r'<[^>]+>').sub('', new_html)
        return result if '\n' in new_html else " ".join(result.split())

    def csv_reader(self, file):
        reader = csv.reader(open(file, encoding='utf_8_sig'))
        new_vacancies = [row for row in reader]
        if len(new_vacancies) == 0:
            print("Пустой файл")
            exit()
        elif len(new_vacancies[1:]) == 0:
            print("Нет данных")
            exit()
        else:
            return new_vacancies[0], new_vacancies[1:]

    def csv_filer(self, headers, vacancies):
        vacancies_list = list(filter(lambda vac: (len(vac) == len(headers) and vac.count('') == 0), vacancies))
        vacanies_dictionary = [dict(zip(headers, map(self.delete_html, vac))) for vac in vacancies_list]
        return vacanies_dictionary

class Salary:
    def __init__(self, salary_from, salary_to, salary_currency):
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_currency = salary_currency

    def to_rub(self, new_salary: float) -> float:
        return new_salary * currency_to_rub[self.salary_currency]

class Vacancy:
    def __init__(self, dictionary):
        self.name = dictionary['name']
        self.salary = Salary(dictionary['salary_from'], dictionary['salary_to'], dictionary['salary_currency'])
        self.area_name = dictionary['area_name']
        self.published_at = dictionary['published_at']

class Report:
    def __init__(self, years_salary, years_vacs_count, prof_years_salary, prof_years_vacs_count, city_salary,
                 city_vacs_rate):
        self.years_salary = years_salary
        self.years_vacs_count = years_vacs_count
        self.prof_years_salary = prof_years_salary
        self.prof_years_vacs_count = prof_years_vacs_count
        self.city_salary = city_salary
        self.city_vacs_rate = city_vacs_rate

    def generate_excel(self):
        new_workbook = Workbook()
        new_workbook.remove(new_workbook.active)
        years_sheet = new_workbook.create_sheet('Статистика по годам')
        city_sheet = new_workbook.create_sheet('Статистика по городам')

        def get_style(sheet):
            for column in sheet.columns:
                new_length = 0
                for col in column:
                    new_side = Side(style="thin", color="000000")
                    col.border = Border(left=new_side, right=new_side, top=new_side, bottom=new_side)
                    if col.value is not None:
                        new_length = max(len(str(col.value)), new_length)
                sheet.column_dimensions[get_column_letter(column[0].column)].width = new_length + 2

        years_sheet_columns = ['Год', 'Средняя зарплата', 'Средняя зарплата - Программист', 'Количество вакансий',
                               'Количество вакансий - Программист']
        for index, column_name in enumerate(years_sheet_columns):
            years_sheet.cell(row=1, column=index + 1, value=column_name).font = Font(bold=True)

        for years in self.years_salary.keys():
            years_sheet.append([years, self.years_salary[years], self.prof_years_salary[years],
                                self.years_vacs_count[years], self.prof_years_vacs_count[years]])
        get_style(years_sheet)

        city_sheet_columns = ['Город', 'Уровень зарплат', ' ', 'Город', 'Доля вакансий']
        for index, column_name in enumerate(city_sheet_columns):
            city_sheet.cell(row=1, column=index + 1, value=column_name).font = Font(bold=True)

        for index, keys in enumerate(self.city_salary.keys()):
            city_sheet.append([keys, self.city_salary[keys], None, list(self.city_vacs_rate.keys())[index],
                               self.city_vacs_rate[list(self.city_vacs_rate.keys())[index]]])
        get_style(city_sheet)

        for cell in city_sheet['E']:
            cell.number_format = FORMAT_PERCENTAGE_00
        new_workbook.save('report.xlsx')


currency_to_rub = {"AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76,
                   "KZT": 0.13, "RUR": 1, "UAH": 1.64, "USD": 60.66, "UZS": 0.0055}

def get_data(date):
    new_date = datetime.datetime.strptime(date, '%Y-%m-%dT%H:%M:%S%z')
    return int(new_date.strftime('%Y'))

def get_statistic(result, index, new_message, slice=0, reverse=False):
    slice = len(result) if slice == 0 else slice
    statistic = dict(sorted(result, key=lambda item: item[index], reverse=reverse)[:slice])
    print(f'{new_message}{str(statistic)}')
    return statistic

def get_salary_statistic(vacs_list: List[Vacancy], fields, vac_name: str = ''):
    statistic_result = {}
    for vac in vacs_list:
        if vac.__getattribute__(fields) not in statistic_result.keys():
            statistic_result[vac.__getattribute__(fields)] = []
    if vac_name != '':
        vacs_list = list(filter(lambda item: vac_name in item.name, vacs_list))

    for vac in vacs_list:
        salary_to = vac.salary.salary_to
        salary_from = vac.salary.salary_from
        statistic_result[vac.__getattribute__(fields)].append(
            vac.salary.to_rub(float(salary_from) + float(salary_to)) / 2)

    for key in statistic_result.keys():
        statistic_result[key] = int(sum(statistic_result[key]) // len(statistic_result[key])) if len(
            statistic_result[key]) != 0 else 0

    return statistic_result

def get_vacancies_statistic(vacs_list: List[Vacancy], fields, vac_name: str = ''):
    statistic_result = {}
    for vac in vacs_list:
        if vac.__getattribute__(fields) not in statistic_result.keys():
            statistic_result[vac.__getattribute__(fields)] = 0

    if vac_name != '':
        vacs_list = list(filter(lambda item: vac_name in item.name, vacs_list))
    for vac in vacs_list:
        statistic_result[vac.__getattribute__(fields)] += 1

    if fields == 'area_name':
        for key in statistic_result.keys():
            statistic_result[key] = round(statistic_result[key] / len(new_data.vacancies), 4)

    return statistic_result

file = input("Введите название файла: ")
vacancy = input("Введите название профессии: ")
new_data = DataSet(file)
new_dict = {}

for vacs in new_data.vacancies:
    vacs.published_at = get_data(vacs.published_at)
    if vacs.area_name not in new_dict.keys():
        new_dict[vacs.area_name] = 0
    new_dict[vacs.area_name] += 1

needed_vacs = list(filter(lambda x: int(len(new_data.vacancies) * 0.01) <= new_dict[x.area_name], new_data.vacancies))
years_salary = get_statistic(get_salary_statistic(new_data.vacancies, 'published_at').items(), 0,
                             'Динамика уровня зарплат по годам: ')
years_vacs_count = get_statistic(get_vacancies_statistic(new_data.vacancies, 'published_at').items(), 0,
                                 'Динамика количества вакансий по годам: ')
prof_years_salary = get_statistic(get_salary_statistic(new_data.vacancies, 'published_at', vacancy).items(), 0,
                                  'Динамика уровня зарплат по годам для выбранной профессии: ')
prof_years_vacs_count = get_statistic(get_vacancies_statistic(new_data.vacancies, 'published_at', vacancy).items(), 0,
                                      'Динамика количества вакансий по годам для выбранной профессии: ')
city_salary = get_statistic(get_salary_statistic(needed_vacs, 'area_name').items(), 1,
                            'Уровень зарплат по городам (в порядке убывания): ', 10, True)
city_vacs_rate = get_statistic(get_vacancies_statistic(needed_vacs, 'area_name').items(), 1,
                               'Доля вакансий по городам (в порядке убывания): ', 10, True)

Report(years_salary, years_vacs_count, prof_years_salary, prof_years_vacs_count, city_salary,
       city_vacs_rate).generate_excel()
